# ==============================================================================
# 1. FERRAMENTA: SELENIUM WEBDRIVER
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: Simula um humano a usar o navegador (clica, escreve, faz scroll).
- Quando usar: Sites que precisam de login, botões que têm de ser clicados 
  ou sites "modernos" que carregam conteúdo enquanto navegas.
- Prós: Interação total com qualquer site.
- Contras: Mais pesado para o computador e mais lento que o BeautifulSoup.
"""
from selenium import webdriver
from selenium.webdriver.chrome.service import Service # Nova importação útil
from webdriver_manager.chrome import ChromeDriverManager # Recomendado instalar
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

# 1. Configuração simplificada
chrome_options = Options()
# chrome_options.add_argument("--headless") # Opcional: roda sem abrir a janela

# 2. Inicialização (Usando o ChromeDriverManager para não ter erro de versão)
driver = webdriver.Chrome(options=chrome_options)

try:
    # --- COMANDOS PRINCIPAIS ---
    driver.get("URL_DO_SITE")                # Abre o site
    time.sleep(2)                            # Pausa de segurança (segundos)
    
    # Encontrar elementos:
    # driver.find_element(By.ID, "id")       # Pelo ID único
    # driver.find_element(By.CSS_SELECTOR, "seletor") # Pela classe ou hierarquia
    
    # Interagir:
    # elemento.click()                       # Clica no botão/link
    # elemento.send_keys("Texto" + Keys.ENTER)# Escreve e prime Enter
    # dado = elemento.text                   # Extrai o texto visível
    
finally:
    driver.quit() # Fecha sempre o navegador para não gastar memória

# ------------------------------------------------------------------------------

# ==============================================================================
# 2. FERRAMENTA: BEAUTIFULSOUP + REQUESTS
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: Descarrega o código HTML do site e lê a informação como um texto.
- Quando usar: Extração rápida de dados em sites simples (notícias, tabelas, 
  preços que já estão na página sem precisar de cliques).
- Prós: Ultra rápido e consome quase nada de memória.
- Contras: Não consegue clicar em botões nem fazer login.
"""
import requests
from bs4 import BeautifulSoup

url = "URL_DO_SITE"

# --- COMANDOS PRINCIPAIS ---
res = requests.get(url)                  # Descarrega o site
res.raise_for_status()                   # Dá erro se o site estiver em baixo

sopa = BeautifulSoup(res.text, 'html.parser') # Transforma em "objeto de leitura"

# Encontrar dados:
# item = sopa.select("seletor_css")       # Retorna uma LISTA com todos os achados
# texto = item[0].getText()               # Limpa as tags e fica só com o texto
# link = item[0].get('href')              # Pega o link escondido num elemento

# ------------------------------------------------------------------------------

# ==============================================================================
# 3. FERRAMENTA: OPENPYXL (EXCEL)
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: Cria, lê e edita ficheiros .xlsx (Excel).
- Quando usar: Guardar os dados que recolheste da internet em tabelas, 
  fazer relatórios automáticos e cálculos matemáticos.
- Prós: Cria ficheiros profissionais que qualquer pessoa consegue abrir.
- Contras: Não podes correr o código se o ficheiro Excel estiver aberto no PC.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- COMANDOS PRINCIPAIS ---
wb = openpyxl.Workbook()                 # Cria um livro novo
aba = wb.active                          # Seleciona a folha principal
aba.title = "Dados"                      # Muda o nome da aba

# Escrever e Calcular:
# aba['A1'] = "Texto"                    # Escreve numa célula
# aba.append(["A", "B", "C"])            # Adiciona uma linha completa
# aba['C2'] = "=A2*B2"                   # Insere fórmula de Excel (Multiplicação)
# aba['D1'] = "=SUM(D2:D10)"             # Insere fórmula de Soma

# Formatação (Visual):
# aba['A1'].font = Font(bold=True)       # Texto em Negrito
# aba.merge_cells('A1:D1')               # Mescla células para títulos

wb.save("NOME_DO_FICHEIRO.xlsx")         # Guarda o trabalho no disco

# ------------------------------------------------------------------------------

# ==============================================================================
# 4. FERRAMENTA: EXPRESSÕES REGULARES (RE)
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: Busca padrões específicos dentro de textos (como procurar uma 
  agulha no palheiro).
- Quando usar: Extrair números de telefone, e-mails, datas ou CPFs de um texto
  bagunçado que você pegou com o Selenium ou BeautifulSoup.
- Prós: Encontra informações que não têm um local fixo na página.
- Contras: A escrita dos padrões (re.compile) pode ser confusa no início.
"""
import re

texto_alvo = "Contacto: leonardo@email.com | Tel: 912-345-678"

# --- COMANDOS PRINCIPAIS ---
# padrao = re.compile(r'\d{3}-\d{3}-\d{3}')  # Cria o padrão (Ex: Telefone)
# resultado = padrao.search(texto_alvo)        # Acha a primeira ocorrência
# todos = padrao.findall(texto_alvo)          # Retorna uma LISTA com todos os achados

# ------------------------------------------------------------------------------

# ==============================================================================
# 5. FERRAMENTA: OS & PATHLIB (ARQUIVOS E PASTAS)
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: Cria pastas, apaga arquivos, lista documentos e organiza o computador.
- Quando usar: Organizar os relatórios de Excel em pastas por data ou verificar 
  se um arquivo já existe antes de tentar abrir.
- Prós: Essencial para rodar scripts no Linux sem dar erro de "Pasta não encontrada".
- Contras: Cuidado ao deletar arquivos, o Python não manda para a lixeira!
"""
import os
from pathlib import Path

# --- COMANDOS PRINCIPAIS ---
# print(os.getcwd())                         # Mostra onde o seu script está rodando
# os.makedirs('Relatorios', exist_ok=True)   # Cria uma pasta nova com segurança
# arquivo = Path('dados.xlsx').exists()      # Verifica se o arquivo existe (True/False)
# lista = os.listdir('.')                    # Lista tudo o que tem na pasta atual

# ------------------------------------------------------------------------------

# ==============================================================================
# 6. FERRAMENTA: TRY / EXCEPT (TRATAMENTO DE ERROS)
# ==============================================================================
"""
DICIONÁRIO DE USO:
- O que faz: É o "colete à prova de balas" do código. Se algo der errado, 
  o programa não "crasha" nem para de funcionar.
- Quando usar: Em TODOS os seus trabalhos de automação (Web e Excel).
- Prós: Permite que o robô continue tentando ou feche os arquivos corretamente 
  mesmo que ocorra um erro de internet ou digitação.
- Contras: Se esconder o erro demais, você pode não saber o que consertar.
"""

# --- ESTRUTURA PRINCIPAL ---
try:
    # bloco de código perigoso (Selenium, Abrir arquivo, etc)
    # x = 1 / 0 
    pass
except Exception as e:
    print(f"ALERTA: Ocorreu o erro -> {e}")
finally:
    # Este bloco SEMPRE roda, serve para limpar a bagunça
    # Ex: driver.quit() ou wb.save()
    pass

# ------------------------------------------------------------------------------